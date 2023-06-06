import argparse
import os
from PIL import Image
from typing import Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import DEFAULT_COLUMN_WIDTH

Pixel = Tuple[int, int, int]
COL_WIDTH = DEFAULT_COLUMN_WIDTH / 3


def create_excel(img, path: str) -> None:
    wb = Workbook()
    ws = wb.active

    for i in range(1, img.height):
        for j in range(1, img.width):
            col_letter = get_column_letter(j)
            if i == 1:
                ws.column_dimensions[col_letter].width = COL_WIDTH
            pixel = img.getpixel((j, i))
            hex: str = "%02x%02x%02x" % (pixel[0], pixel[1], pixel[2])
            ws[col_letter + str(i)].fill = PatternFill(fill_type="solid",
                                                       start_color=hex,
                                                       end_color=hex)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Creates an Excel sheet and copies an image into it.")
    parser.add_argument("image_path", help="Path to input image.")
    args: Dict[str, str] = vars(parser.parse_args())

    name: str = os.path.splitext(os.path.basename(args["image_path"]))[0] + ".xlsx"

    try:
        with Image.open(args["image_path"]) as img:
            rgb_img = img.convert("RGB")
    except FileNotFoundError:
        print(f"Image with name \"{args['image_path']}\" doesn't exist in the current diectory.")
        return

    print("Copying image ...\n")

    create_excel(rgb_img, name)

    print(f"Created file \"{name}\" in the current directory.")
    print("Image copied succesfully!")


if __name__ == "__main__":
    main()
